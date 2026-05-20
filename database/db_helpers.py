from core.paths import DATA_DIR, TRADES_PATH, UPDATES_PATH
from utils.column_mapper import DEFAULT_HEADERS, map_row_to_trade
from utils.logger import setup_logger
from openpyxl import load_workbook, Workbook
from datetime import datetime
import threading
import os

logger = setup_logger("DBHelpers")

_cache = {
    "trades": {"data": None, "mtime": None},
    "updates": {"data": None, "mtime": None},
}
_cache_lock = threading.Lock()


def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _load_wb(path, headers: list):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save(path)
    return wb


def _save_workbook(wb, path):
    wb.save(path)


def invalidate_cache(*keys):
    with _cache_lock:
        for key in keys:
            if key in _cache:
                _cache[key]["data"] = None
                _cache[key]["mtime"] = None
        logger.debug(f"Cache invalidated for: {keys}")


def _get_cached_rows(path, headers, is_trades=True):
    cache_key = "trades" if is_trades else "updates"
    mtime = os.path.getmtime(path) if path.exists() else None

    with _cache_lock:
        cached = _cache[cache_key]
        if cached["data"] is not None and cached["mtime"] == mtime:
            return cached["data"]

    wb = _load_wb(path, headers)
    ws = wb.active
    rows = _wb_to_dicts(ws, headers, is_trades=is_trades)

    with _cache_lock:
        _cache[cache_key]["data"] = rows
        _cache[cache_key]["mtime"] = mtime

    return rows


def _wb_to_dicts(ws, headers=None, is_trades=True):
    rows = []
    sheet_headers = [c.value for c in ws[1]]
    if not any(sheet_headers):
        sheet_headers = headers if headers else DEFAULT_HEADERS()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        processed = []
        for v in row:
            if isinstance(v, datetime):
                processed.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                processed.append(v)
        row_dict = dict(zip(sheet_headers, processed))
        if is_trades:
            rows.append(map_row_to_trade(row_dict))
        else:
            rows.append(row_dict)
    return rows