"""
db_manager.py  –  xlsx-based local storage (replaces SQLite)

Two files are managed:
  data/trades.xlsx        – one row per trade
  data/trade_updates.xlsx – one row per update event

Settings (get_setting / set_setting) are stored in data/settings.json.
"""

import json
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook, Workbook
from utils.logger import setup_logger

logger = setup_logger("DBManager")

DATA_DIR      = Path(__file__).resolve().parent.parent / "data"
TRADES_PATH   = DATA_DIR / "trades.xlsx"
UPDATES_PATH  = DATA_DIR / "trade_updates.xlsx"
SETTINGS_PATH = DATA_DIR / "settings.json"

# ─── Column definitions ────────────────────────────────────────────────────────

from utils.column_mapper import DEFAULT_HEADERS, map_row_to_trade, map_trade_to_columns

TRADES_HEADERS = DEFAULT_HEADERS

UPDATES_HEADERS = [
    "id", "trade_id", "update_type", "details",
    "old_value", "new_value", "created_at",
]

# ─── Internal helpers ──────────────────────────────────────────────────────────

def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

def _load_wb(path: Path, headers: list) -> openpyxl.Workbook:
    """Load workbook or create it with a header row if missing."""
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save(path)
    return wb

def _wb_to_dicts(ws, headers: list = None) -> list[dict]:
    """Read all data rows from a worksheet into a list of dicts. Maps them back to internal names."""
    rows = []
    
    # Actually fetch the headers from the sheet to be dynamic
    sheet_headers = [c.value for c in ws[1]]
    if not any(sheet_headers):
        sheet_headers = headers if headers else DEFAULT_HEADERS
        
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        # Convert datetime objects to ISO strings for consistency
        processed = []
        for v in row:
            if isinstance(v, datetime):
                processed.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                processed.append(v)
        row_dict = dict(zip(sheet_headers, processed))
        # If it's the trades sheet, map it back correctly:
        if "update_type" not in sheet_headers: # hacky check to differentiate trades vs trade_updates sheet
            rows.append(map_row_to_trade(row_dict))
        else:
            rows.append(row_dict)
    return rows


def _save_trades(wb: openpyxl.Workbook):
    wb.save(TRADES_PATH)

def _save_updates(wb: openpyxl.Workbook):
    wb.save(UPDATES_PATH)

def _next_id(ws) -> int:
    """Return max existing id + 1 (1 if sheet is empty beyond header)."""
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, int) and val > max_id:
            max_id = val
    return max_id + 1

def generate_trade_code() -> str:
    date_str = datetime.now().strftime("%Y%m%d")
    suffix = uuid.uuid4().hex[:6].upper()
    return f"TRD-{date_str}-{suffix}"

def _unique_trade_code(all_rows: list[dict]) -> str:
    existing = {r.get("trade_code") for r in all_rows}
    for _ in range(10):
        code = generate_trade_code()
        if code not in existing:
            return code
    raise RuntimeError("Unable to generate a unique trade code after 10 attempts.")

# ─── Public API ────────────────────────────────────────────────────────────────

def init_db():
    """Create xlsx files with header rows if they don't already exist."""
    try:
        _ensure_data_dir()
        _load_wb(TRADES_PATH,  TRADES_HEADERS)
        _load_wb(UPDATES_PATH, UPDATES_HEADERS)
        logger.info("Data store initialised (xlsx).")
    except Exception as e:
        logger.error(f"Error initialising data store: {e}", exc_info=True)
        raise


def insert_trade(trade_data: dict) -> int:
    """Append a new trade row to trades.xlsx. Returns the new integer trade id.
    Also mutates trade_data in-place with id, trade_code, created_at, updated_at."""
    try:
        _ensure_data_dir()
        # Fetch existing headers
        try:
            wb = load_workbook(TRADES_PATH)
            ws = wb.active
            headers = [c.value for c in ws[1]]
        except Exception:
            wb = _load_wb(TRADES_PATH, TRADES_HEADERS)
            ws = wb.active
            headers = TRADES_HEADERS
            
        all_rows = _wb_to_dicts(ws, headers)
        trade_code = _unique_trade_code(all_rows)
        trade_id   = _next_id(ws)
        now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Mutate caller's dict so mapping catches them
        trade_data["id"]         = trade_id
        trade_data["trade_code"] = trade_code
        trade_data["created_at"] = now
        trade_data["updated_at"] = now

        # Map to columns
        row = map_trade_to_columns(trade_data, headers, is_google_sheets=False)

        ws.append(row)
        _save_trades(wb)

        logger.info(f"Inserted trade ID {trade_id} code {trade_code} ({trade_data.get('stock_name')})")
        return trade_id
    except Exception as e:
        logger.error(f"Error inserting trade: {e}", exc_info=True)
        raise


def get_trade(trade_id: int) -> dict | None:
    """Return a single trade dict by id, or None."""
    try:
        wb = _load_wb(TRADES_PATH, TRADES_HEADERS)
        ws = wb.active
        for row in _wb_to_dicts(ws, TRADES_HEADERS):
            if row.get("id") == trade_id:
                return row
        return None
    except Exception as e:
        logger.error(f"Error fetching trade ID {trade_id}: {e}", exc_info=True)
        return None


def get_all_trades(filters: dict | None = None) -> list[dict]:
    """Return all trades as list of dicts, optionally filtered, newest first."""
    try:
        wb = _load_wb(TRADES_PATH, TRADES_HEADERS)
        ws = wb.active
        rows = _wb_to_dicts(ws, TRADES_HEADERS)

        if filters:
            if filters.get("status"):
                rows = [r for r in rows if r.get("status") == filters["status"]]
            if filters.get("segment"):
                rows = [r for r in rows if r.get("segment") == filters["segment"]]
            if filters.get("search"):
                term = filters["search"].lower()
                rows = [r for r in rows if term in str(r.get("stock_name", "")).lower()]

        # Sort newest first by created_at (string ISO format sorts correctly)
        rows.sort(key=lambda r: r.get("created_at") or "", reverse=True)
        return rows
    except Exception as e:
        logger.error(f"Error fetching all trades: {e}", exc_info=True)
        return []


def update_trade(trade_id: int, fields: dict) -> bool:
    """Update columns for an existing trade row in trades.xlsx."""
    try:
        # Fetch old trade and merge
        trade_data = get_trade(trade_id)
        if not trade_data:
            logger.warning(f"Trade ID {trade_id} not found for update.")
            return False
            
        trade_data.update(fields)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        trade_data["updated_at"] = now
        
        wb = _load_wb(TRADES_PATH, TRADES_HEADERS)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        if not any(headers):
            headers = TRADES_HEADERS

        # Map to columns
        row_arr = map_trade_to_columns(trade_data, headers, is_google_sheets=False)

        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == trade_id:  # Column A is ALWAYS ID per map_trade_to_columns logic
                for col_idx, val in enumerate(row_arr, start=1):
                    ws.cell(row=row[0].row, column=col_idx, value=val)
                updated = True
                break

        if updated:
            _save_trades(wb)
            logger.info(f"Updated trade ID {trade_id} with {fields}")
        return updated
    except Exception as e:
        logger.error(f"Error updating trade ID {trade_id}: {e}", exc_info=True)
        return False


def insert_trade_update(update_data: dict) -> int:
    """Append an update event row to trade_updates.xlsx. Returns new id."""
    try:
        _ensure_data_dir()
        wb = _load_wb(UPDATES_PATH, UPDATES_HEADERS)
        ws = wb.active

        update_id = _next_id(ws)
        now       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        old_val = update_data.get("old_value")
        new_val = update_data.get("new_value")
        row = [
            update_id,
            update_data["trade_id"],
            update_data["update_type"],
            update_data.get("details", ""),
            json.dumps(old_val) if old_val is not None else "",
            json.dumps(new_val) if new_val is not None else "",
            now,
        ]
        ws.append(row)
        _save_updates(wb)

        logger.info(f"Inserted trade update for trade ID {update_data['trade_id']} (Type: {update_data['update_type']})")
        return update_id
    except Exception as e:
        logger.error(f"Error inserting trade update: {e}", exc_info=True)
        raise


def get_trade_updates(trade_id: int) -> list[dict]:
    """Return all update events for a given trade id, newest first."""
    try:
        wb = _load_wb(UPDATES_PATH, UPDATES_HEADERS)
        ws = wb.active
        rows = _wb_to_dicts(ws, UPDATES_HEADERS)
        filtered = [r for r in rows if r.get("trade_id") == trade_id]
        filtered.sort(key=lambda r: r.get("created_at") or "", reverse=True)
        return filtered
    except Exception as e:
        logger.error(f"Error fetching trade updates for trade ID {trade_id}: {e}", exc_info=True)
        return []


# ─── Settings (simple JSON sidecar) ───────────────────────────────────────────

def _load_settings() -> dict:
    if SETTINGS_PATH.exists():
        try:
            with open(SETTINGS_PATH) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_settings(data: dict):
    _ensure_data_dir()
    with open(SETTINGS_PATH, "w") as f:
        json.dump(data, f, indent=2)

def get_setting(key: str) -> str | None:
    try:
        return _load_settings().get(key)
    except Exception as e:
        logger.error(f"Error fetching setting {key}: {e}", exc_info=True)
        return None

def set_setting(key: str, value: str):
    try:
        data = _load_settings()
        data[key] = value
        _save_settings(data)
    except Exception as e:
        logger.error(f"Error setting {key} to {value}: {e}", exc_info=True)
