from core.paths import DATA_DIR, TRADES_PATH
from database.db_helpers import (
    _ensure_data_dir, _load_wb, _save_workbook, _wb_to_dicts,
    _get_cached_rows, invalidate_cache,
)
from utils.column_mapper import DEFAULT_HEADERS, map_row_to_trade, map_trade_to_columns
from utils.logger import setup_logger
from openpyxl import load_workbook
from datetime import datetime
import uuid

logger = setup_logger("TradesDB")

TRADES_HEADERS = DEFAULT_HEADERS()


def _ensure_schema_columns():
    try:
        if not TRADES_PATH.exists():
            return
        wb = load_workbook(TRADES_PATH)
        ws = wb.active
        existing = [c.value for c in ws[1]]
        added = False
        for header in TRADES_HEADERS:
            if header not in existing:
                ws.cell(row=1, column=len(existing) + 1, value=header)
                existing.append(header)
                added = True
        if added:
            _save_workbook(wb, TRADES_PATH)
            invalidate_cache("trades")
            logger.info("Added missing schema columns to trades.xlsx")
    except Exception as e:
        logger.error(f"Error ensuring schema columns: {e}", exc_info=True)


def get_trades_headers():
    return TRADES_HEADERS


def insert_trade(trade_data: dict) -> tuple[str, dict]:
    try:
        _ensure_data_dir()
        _ensure_schema_columns()
        try:
            wb = load_workbook(TRADES_PATH)
            ws = wb.active
            headers = [c.value for c in ws[1]]
        except Exception:
            wb = _load_wb(TRADES_PATH, TRADES_HEADERS)
            ws = wb.active
            headers = TRADES_HEADERS

        all_rows = _wb_to_dicts(ws, headers, is_trades=True)
        existing_codes = {r.get("trade_code") for r in all_rows}
        trade_code = _unique_trade_code(existing_codes)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        enriched = {**trade_data, "trade_code": trade_code, "created_at": now, "updated_at": now}

        row = map_trade_to_columns(enriched, headers, is_google_sheets=False)
        ws.append(row)
        _save_workbook(wb, TRADES_PATH)
        invalidate_cache("trades")

        logger.info(f"Inserted trade code {trade_code} ({enriched.get('stock_name')})")
        return trade_code, enriched
    except Exception as e:
        logger.error(f"Error inserting trade: {e}", exc_info=True)
        raise


def get_trade(trade_code: str) -> dict | None:
    try:
        rows = _get_cached_rows(TRADES_PATH, TRADES_HEADERS, is_trades=True)
        for row in rows:
            if row.get("trade_code") == trade_code:
                return row
        return None
    except Exception as e:
        logger.error(f"Error fetching trade code {trade_code}: {e}", exc_info=True)
        return None


def get_all_trades(filters: dict | None = None) -> list[dict]:
    try:
        rows = _get_cached_rows(TRADES_PATH, TRADES_HEADERS, is_trades=True)

        if filters:
            if filters.get("status"):
                rows = [r for r in rows if r.get("status") == filters["status"]]
            if filters.get("segment"):
                rows = [r for r in rows if r.get("segment") == filters["segment"]]
            if filters.get("search"):
                term = filters["search"].lower()
                rows = [r for r in rows if term in str(r.get("stock_name", "")).lower()]

        rows.sort(key=lambda r: r.get("created_at") or "", reverse=True)
        return rows
    except Exception as e:
        logger.error(f"Error fetching all trades: {e}", exc_info=True)
        return []


def update_trade(trade_code: str, fields: dict) -> bool:
    if not trade_code or trade_code == "?":
        raise ValueError("Strict Check Failed: trade_code is required to update a trade.")

    try:
        _ensure_schema_columns()
        trade_data = get_trade(trade_code)
        if not trade_data:
            logger.warning(f"Trade code {trade_code} not found for update.")
            return False

        trade_data.update(fields)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        trade_data["updated_at"] = now

        wb = load_workbook(TRADES_PATH)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        if not any(headers):
            headers = TRADES_HEADERS

        row_arr = map_trade_to_columns(trade_data, headers, is_google_sheets=False)

        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == trade_code:
                for col_idx, val in enumerate(row_arr, start=1):
                    ws.cell(row=row[0].row, column=col_idx, value=val)
                updated = True
                break

        if updated:
            _save_workbook(wb, TRADES_PATH)
            invalidate_cache("trades")
            logger.info(f"Updated trade code {trade_code} with {fields}")
        return updated
    except Exception as e:
        logger.error(f"Error updating trade code {trade_code}: {e}", exc_info=True)
        return False


def generate_trade_code() -> str:
    date_str = datetime.now().strftime("%Y%m%d")
    suffix = uuid.uuid4().hex[:6].upper()
    return f"TRD-{date_str}-{suffix}"


def _unique_trade_code(existing_codes: set) -> str:
    for _ in range(10):
        code = generate_trade_code()
        if code not in existing_codes:
            return code
    raise RuntimeError("Unable to generate a unique trade code after 10 attempts.")